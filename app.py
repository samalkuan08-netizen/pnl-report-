"""
P&L-агент Градекс KZ — веб-приложение.
Загрузка выгрузки 1С -> автоматическое заполнение P&L по шаблону -> готовый Excel + журнал.
Запуск локально:  streamlit run app.py
Деплой: GitHub -> Streamlit Community Cloud (положить рядом template.xlsx).
"""
import io, os, tempfile
from collections import defaultdict
import streamlit as st
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import pnl_agent as eng

st.set_page_config(page_title="P&L-агент Градекс KZ", page_icon="📊", layout="wide")

MONTHS = ['Январь','Февраль','Март','Апрель','Май','Июнь',
          'Июль','Август','Сентябрь','Октябрь','Ноябрь','Декабрь']
MONTH_COL = {m: get_column_letter(3+i) for i,m in enumerate(MONTHS)}   # C..N

st.title("📊 P&L-агент Градекс KZ")
st.caption("Загрузите выгрузку 1С (Анализ счёта с субконто) — агент сам разнесёт доходы и "
           "расходы по дивизионам ТШО / КПО / Бетон / АУП и заполнит P&L по шаблону.")

with st.sidebar:
    st.header("Настройки")
    month = st.selectbox("Месяц отчёта", MONTHS, index=0)
    threshold = st.number_input("Порог новой статьи, ₸", value=300_000, step=50_000,
                                help="Новая статья ≥ порога → отдельная строка; ниже → «Прочие».")
    eng.THRESHOLD = int(threshold)
    st.divider()
    st.subheader("Шаблон")
    tmpl_up = st.file_uploader("Шаблон P&L (.xlsx)", type=['xlsx'], key='tmpl',
                               help="Если в репозитории есть template.xlsx — можно не загружать.")

c1, c2 = st.columns([1,1])
with c1:
    data_up = st.file_uploader("Выгрузка 1С за месяц (.xlsx)", type=['xlsx'], key='data')

def resolve_template():
    if tmpl_up is not None:
        return tmpl_up.getvalue()
    if os.path.exists('template.xlsx'):
        return open('template.xlsx','rb').read()
    return None

run = st.button("▶️ Сформировать отчёт", type="primary", disabled=(data_up is None))

def build_journal(wb, log):
    if 'Журнал сопоставления' in wb.sheetnames: del wb['Журнал сопоставления']
    js = wb.create_sheet('Журнал сопоставления')
    js.append(['Счёт 1С','Статья 1С','Дивизион','Тип','Лист отчёта','Строка отчёта','Сумма ₸'])
    thin=Side(style='thin',color='D0D0D0'); bd=Border(thin,thin,thin,thin)
    for c in js[1]:
        c.font=Font(bold=True,color='FFFFFF',name='Arial',size=10); c.fill=PatternFill('solid',fgColor='2F5496')
        c.alignment=Alignment(horizontal='center',wrap_text=True); c.border=bd
    for row in sorted(log,key=lambda x:(x[4],str(x[5]))):
        js.append([row[0],row[1],row[2],row[3],row[4],str(row[5]),round(float(row[6]),2)])
    for i,w in enumerate([10,46,10,12,20,42,16],1): js.column_dimensions[get_column_letter(i)].width=w
    last=js.max_row
    for r in range(2,last+1):
        for cc in range(1,8):
            cell=js.cell(row=r,column=cc); cell.font=Font(name='Arial',size=10); cell.border=bd
            if cc==7: cell.number_format='#,##0;(#,##0);-'; cell.alignment=Alignment(horizontal='right')
    js.freeze_panes='A2'
    js.cell(row=last+1,column=6,value='ИТОГО').font=Font(bold=True,name='Arial')
    t=js.cell(row=last+1,column=7,value=f'=SUM(G2:G{last})'); t.font=Font(bold=True,name='Arial'); t.number_format='#,##0'

if run:
    tmpl_bytes = resolve_template()
    if tmpl_bytes is None:
        st.error("Не найден шаблон. Загрузите его в боковой панели или положите template.xlsx в репозиторий.")
        st.stop()
    with tempfile.TemporaryDirectory() as td:
        dpath=os.path.join(td,'data.xlsx'); tpath=os.path.join(td,'tmpl.xlsx'); opath=os.path.join(td,'out.xlsx')
        open(dpath,'wb').write(data_up.getvalue()); open(tpath,'wb').write(tmpl_bytes)
        try:
            leaves = eng.parse_1c(dpath)
        except Exception as ex:
            st.error(f"Не удалось разобрать выгрузку 1С: {ex}"); st.stop()
        acc, flags, log = eng.fill_report(tpath, leaves, opath, month_col=MONTH_COL[month])

        # сверка по дивизионам
        ref=defaultdict(float); inc=defaultdict(float)
        for (ac,it,k,d,v) in leaves: (ref if k=='expense' else inc)[d]+=v
        total_exp=sum(ref.values()); total_inc=sum(inc.values())

        st.success(f"Отчёт за {month} сформирован. Разнесено статей: {len(log)}.")
        m1,m2,m3,m4=st.columns(4)
        m1.metric("Доходы, ₸", f"{total_inc:,.0f}")
        m2.metric("Расходы, ₸", f"{total_exp:,.0f}")
        m3.metric("Результат до КПН, ₸", f"{total_inc-total_exp:,.0f}")
        m4.metric("Новых статей (флаги)", len(flags))

        st.subheader("Сверка с 1С (контроль, что ничего не потеряно)")
        import pandas as pd
        rec=pd.DataFrame([
            {'Дивизион':'ТШО','1С, ₸':ref['ТШО']},
            {'Дивизион':'КПО','1С, ₸':ref['КПО']},
            {'Дивизион':'Бетон','1С, ₸':ref['Бетон']},
            {'Дивизион':'АУП','1С, ₸':ref['АУП']},
            {'Дивизион':'ИТОГО расходы','1С, ₸':total_exp},
        ])
        rec['1С, ₸']=rec['1С, ₸'].map(lambda x:f"{x:,.0f}")
        st.dataframe(rec, hide_index=True, use_container_width=True)

        if flags:
            st.subheader("⚠️ Новые статьи на проверку")
            st.caption("Статей нет в шаблоне. ≥ порога — добавлены отдельной строкой/во флаг; ниже — в «Прочие».")
            fdf=pd.DataFrame(flags, columns=['Статья 1С','Дивизион','Сумма ₸','Действие'])
            fdf['Сумма ₸']=fdf['Сумма ₸'].map(lambda x:f"{x:,.0f}")
            st.dataframe(fdf, hide_index=True, use_container_width=True)

        st.subheader("Журнал сопоставления")
        ldf=pd.DataFrame(log, columns=['Счёт','Статья 1С','Дивизион','Тип','Лист','Строка отчёта','Сумма ₸'])
        ldf['Сумма ₸']=ldf['Сумма ₸'].map(lambda x:f"{x:,.0f}")
        st.dataframe(ldf, hide_index=True, use_container_width=True, height=320)

        # журнал в книгу + отдача
        wb=openpyxl.load_workbook(opath); build_journal(wb, log)
        buf=io.BytesIO(); wb.save(buf); buf.seek(0)
        st.download_button("⬇️ Скачать заполненный P&L (.xlsx)", buf,
            file_name=f"GRD_{month}_заполнен.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary")
        st.info("Лист «P&L 2026 all» в файле пересчитается формулами при открытии в Excel/LibreOffice.")
else:
    st.info("Загрузите выгрузку 1С и нажмите «Сформировать отчёт».")
